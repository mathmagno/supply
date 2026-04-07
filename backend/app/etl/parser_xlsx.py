"""
Parser da planilha de cotação usando openpyxl (puro Python, sem pandas).

Estrutura esperada das colunas:
  A: Data  |  B: produto (código)  |  C: marca  |  D: especificacao
  E: Custo planejado  |  F: Último custo
  G...: fornecedor 1, fornecedor 2, ... (quantidade dinâmica)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl


# Índices fixos (1-based, conforme openpyxl)
COL_DATA = 1
COL_PRODUTO = 2
COL_MARCA = 3
COL_ESPECIFICACAO = 4
COL_CUSTO_PLAN = 5
COL_ULTIMO_CUSTO = 6
COL_FORNECEDORES_INICIO = 7  # a partir da coluna G


@dataclass
class LinhaplanilhaResult:
    data_cotacao: str          # "YYYY-MM-DD"
    codigo: str                # produto (coluna B)
    marca: str | None
    descricao: str             # especificacao (coluna D)
    custo_planejado: Decimal | None
    ultimo_custo: Decimal | None
    precos_fornecedores: dict[int, Decimal] = field(default_factory=dict)
    nomes_fornecedores: dict[int, str] = field(default_factory=dict)  # posicao → nome
    erros: list[str] = field(default_factory=list)


def _para_decimal(valor) -> Decimal | None:
    if valor is None or valor == "":
        return None
    try:
        return Decimal(str(valor)).quantize(Decimal("0.0001"))
    except InvalidOperation:
        return None


def _para_data(valor) -> str:
    if isinstance(valor, (datetime, date)):
        d = valor if isinstance(valor, date) else valor.date()
        return d.strftime("%Y-%m-%d")
    if isinstance(valor, str) and valor.strip():
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(valor.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return datetime.today().strftime("%Y-%m-%d")


def _detectar_colunas_fornecedor(header_row: tuple) -> dict[int, tuple[int, str]]:
    """
    Detecta dinamicamente colunas de fornecedor a partir do cabeçalho.
    Retorna {posicao: (col_idx_1based, nome_coluna)}.
    Considera fornecedor qualquer coluna a partir de COL_FORNECEDORES_INICIO
    que tenha valor não-vazio no cabeçalho.
    """
    colunas: dict[int, tuple[int, str]] = {}
    posicao = 1
    for col_idx_0 in range(COL_FORNECEDORES_INICIO - 1, len(header_row)):
        nome = str(header_row[col_idx_0] or "").strip()
        if nome:
            colunas[posicao] = (col_idx_0 + 1, nome)  # converte para 1-based
            posicao += 1
    return colunas


def parsear_xlsx(caminho: str | Path) -> list[LinhaplanilhaResult]:
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb.active

    # Lê o cabeçalho para detectar colunas de fornecedor dinamicamente
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    colunas_fornecedor = _detectar_colunas_fornecedor(header_row)
    nomes_fornecedores = {pos: nome for pos, (_, nome) in colunas_fornecedor.items()}

    resultados: list[LinhaplanilhaResult] = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def cel(col_idx: int):
            try:
                v = row[col_idx - 1]
                return v if v is not None else None
            except IndexError:
                return None

        codigo = str(cel(COL_PRODUTO) or "").strip()
        if not codigo or codigo.lower() == "none":
            continue

        erros: list[str] = []

        data_raw = cel(COL_DATA)
        try:
            data_cotacao = _para_data(data_raw)
        except Exception:
            data_cotacao = datetime.today().strftime("%Y-%m-%d")
            erros.append(f"Linha {idx}: data inválida, usando data de hoje.")

        precos: dict[int, Decimal] = {}
        for posicao, (col_idx, _) in colunas_fornecedor.items():
            dec = _para_decimal(cel(col_idx))
            if dec is not None:
                precos[posicao] = dec

        resultado = LinhaplanilhaResult(
            data_cotacao=data_cotacao,
            codigo=codigo,
            marca=str(cel(COL_MARCA) or "").strip() or None,
            descricao=str(cel(COL_ESPECIFICACAO) or "").strip(),
            custo_planejado=_para_decimal(cel(COL_CUSTO_PLAN)),
            ultimo_custo=_para_decimal(cel(COL_ULTIMO_CUSTO)),
            precos_fornecedores=precos,
            nomes_fornecedores=nomes_fornecedores,
            erros=erros,
        )
        resultados.append(resultado)

    wb.close()
    return resultados
