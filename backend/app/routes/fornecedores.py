import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.auth import get_usuario_atual
from app.models.fornecedor import Fornecedor
from app.models.pedido import Pedido, StatusPedido
from app.schemas.fornecedor import FornecedorCreate, FornecedorUpdate, FornecedorOut

router = APIRouter()


@router.get("/modelo-xlsx")
def modelo_xlsx(
    segmento: str = Query(..., description="Segmento dos fornecedores"),
    db: Session = Depends(get_db),
    _=Depends(get_usuario_atual),
):
    """Gera planilha .xlsx modelo com os fornecedores do segmento como cabeçalho."""
    fornecedores = (
        db.query(Fornecedor)
        .filter(Fornecedor.segmento == segmento, Fornecedor.ativo == 1)
        .order_by(Fornecedor.nome)
        .all()
    )
    if not fornecedores:
        raise HTTPException(status_code=404, detail=f"Nenhum fornecedor ativo no segmento '{segmento}'.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotação"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    fornec_fill = PatternFill("solid", fgColor="2563EB")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cabeçalhos fixos
    cabecalhos_fixos = ["Data", "produto", "marca", "especificacao", "Custo planejado", "Último custo"]
    cabecalhos = cabecalhos_fixos + [f.nome for f in fornecedores]

    for col, titulo in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        if col <= len(cabecalhos_fixos):
            cell.fill = header_fill
        else:
            cell.fill = fornec_fill

    # Linha de exemplo
    ws.cell(row=2, column=1, value=date.today().strftime("%d/%m/%Y"))
    ws.cell(row=2, column=2, value="codigo01")
    ws.cell(row=2, column=3, value="marca")
    ws.cell(row=2, column=4, value="Descrição do produto")
    ws.cell(row=2, column=5, value=0.00)
    ws.cell(row=2, column=6, value=0.00)
    for col in range(len(cabecalhos_fixos) + 1, len(cabecalhos) + 1):
        ws.cell(row=2, column=col, value=0.00)

    # Larguras das colunas
    larguras = [14, 12, 14, 35, 16, 14] + [18] * len(fornecedores)
    for col, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = largura

    ws.row_dimensions[1].height = 22

    # Salva em memória
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = f"cotacao_{segmento.lower().replace(' ', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arquivo}"'},
    )


@router.get("/segmentos")
def listar_segmentos(db: Session = Depends(get_db), _=Depends(get_usuario_atual)):
    """Retorna lista de segmentos distintos cadastrados."""
    rows = (
        db.query(Fornecedor.segmento)
        .filter(Fornecedor.segmento.isnot(None), Fornecedor.ativo == 1)
        .distinct()
        .order_by(Fornecedor.segmento)
        .all()
    )
    return [r[0] for r in rows]


def _calcular_lead_times(db: Session) -> dict[int, int]:
    """Calcula a média real de dias (data_pedido → data_recebimento) por fornecedor."""
    pedidos = (
        db.query(Pedido.fornecedor_id, Pedido.data_pedido, Pedido.data_recebimento)
        .filter(
            Pedido.status == StatusPedido.RECEBIDO,
            Pedido.data_pedido.isnot(None),
            Pedido.data_recebimento.isnot(None),
        )
        .all()
    )
    totais: dict[int, list[int]] = {}
    for fid, dp, dr in pedidos:
        dias = (dr - dp).days
        if dias >= 0:
            totais.setdefault(fid, []).append(dias)
    return {fid: round(sum(v) / len(v)) for fid, v in totais.items()}


@router.get("/", response_model=list[FornecedorOut])
def listar(
    apenas_ativos: bool = Query(True),
    db: Session = Depends(get_db),
    _=Depends(get_usuario_atual),
):
    q = db.query(Fornecedor)
    if apenas_ativos:
        q = q.filter(Fornecedor.ativo == 1)
    fornecedores = q.order_by(Fornecedor.nome).all()

    lead_times = _calcular_lead_times(db)

    resultado = []
    for f in fornecedores:
        out = FornecedorOut.model_validate(f)
        out.lead_time_calculado = lead_times.get(f.id)
        resultado.append(out)
    return resultado


@router.post("/", response_model=FornecedorOut, status_code=201)
def criar(payload: FornecedorCreate, db: Session = Depends(get_db), _=Depends(get_usuario_atual)):
    if db.query(Fornecedor).filter(Fornecedor.nome == payload.nome).first():
        raise HTTPException(status_code=400, detail="Fornecedor já cadastrado.")
    fornecedor = Fornecedor(**payload.model_dump())
    db.add(fornecedor)
    db.commit()
    db.refresh(fornecedor)
    return fornecedor


@router.get("/{id}", response_model=FornecedorOut)
def obter(id: int, db: Session = Depends(get_db), _=Depends(get_usuario_atual)):
    f = db.get(Fornecedor, id)
    if not f:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado.")
    return f


@router.patch("/{id}", response_model=FornecedorOut)
def atualizar(id: int, payload: FornecedorUpdate, db: Session = Depends(get_db), _=Depends(get_usuario_atual)):
    f = db.get(Fornecedor, id)
    if not f:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado.")
    for campo, valor in payload.model_dump(exclude_none=True).items():
        setattr(f, campo, valor)
    db.commit()
    db.refresh(f)
    return f


@router.delete("/{id}", status_code=204)
def remover(id: int, db: Session = Depends(get_db), _=Depends(get_usuario_atual)):
    from app.models.cotacao import Cotacao
    from app.models.pedido import Pedido

    f = db.get(Fornecedor, id)
    if not f:
        raise HTTPException(status_code=404, detail="Fornecedor não encontrado.")

    # Remove cotações vinculadas
    db.query(Cotacao).filter(Cotacao.fornecedor_id == id).delete()
    # Remove pedidos vinculados
    db.query(Pedido).filter(Pedido.fornecedor_id == id).delete()

    db.delete(f)
    db.commit()
